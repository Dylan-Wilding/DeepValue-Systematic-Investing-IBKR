# ==============================================================================
# Date: 2026-02-23
# Author: Gemini 3.1 Pro
# Objective: Pass 1, 2, 3 & 4 - End-to-End Execution + Complete Audit Engine
# Project: Forward P/E Systematic Strategy (Intelligent Local Data Caching)
# ==============================================================================

import pandas as pd
import numpy as np
import yfinance as yf
import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
import os
import json
import datetime
import time
import random
import io
import warnings
import asyncio
import scipy.stats as stats
import matplotlib.pyplot as plt

try:
    from ib_async import IB, Stock, MarketOrder, util
except ImportError:
    print("CRITICAL WARNING: 'ib_async' is missing. Please run 'pip install ib_async'")
    from ib_insync import IB, Stock, MarketOrder, util

try:
    import openpyxl
    from openpyxl.drawing.image import Image as OpenpyxlImage
    from openpyxl.styles import Font
except ImportError:
    print("CRITICAL WARNING: 'openpyxl' is missing. Trade Logging and Audit Reporting require it.")
    print("Please run 'pip install openpyxl'")

# Suppress warnings for cleaner verbose output
warnings.simplefilter(action='ignore', category=FutureWarning)
pd.options.display.float_format = '{:,.2f}'.format

# --- CONFIGURATION ---
TEST_MODE = False
BASE_EQUITY_FALLBACK = 1000000.0
EXCLUDE_REBOUNDS = True
FILENAME_DB = "portfolios_pe_ladders.json"
REAL_FILENAME_DB = "real_ibkr_portfolio.json"
TRADE_LOG_FILE = "trade_log.xlsx"
SCREENER_OUTPUT_FILE = "screener_results.xlsx"
MIN_HOLDINGS = 20
MAX_HOLDINGS = 25

# --- IBKR CONFIGURATION ---
IB_IP = '127.0.0.1'
IB_PORT = 4002
IB_CLIENT_ID = 1
TARGET_ACCOUNT = "DUP807776"

# --- AUDIT CONFIGURATION ---
BENCHMARK_TICKER = "^GSPC"
RISK_FREE_RATE = 0.04
MONTE_CARLO_SIMS = 10000

# --- GLOBAL SESSION LIMITER (For Wikipedia Only) ---
session = requests.Session()
retry = Retry(connect=3, backoff_factor=1.0, status_forcelist=[429, 500, 502, 503, 504])
adapter = HTTPAdapter(max_retries=retry)
session.mount('http://', adapter)
session.mount('https://', adapter)
session.headers.update({
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
})


def log(msg):
    timestamp = datetime.datetime.now().strftime("%H:%M:%S")
    print(f"[{timestamp}] {msg}")


def print_header(title):
    print("\n" + "=" * 120)
    print(f" {title}")
    print("=" * 120)


# ==============================================================================
# STARTUP DASHBOARD ENGINE
# ==============================================================================

def display_startup_dashboard():
    ib_dash = IB()
    try:
        # Use a distinct client ID to prevent blocking the main process
        ib_dash.connect(IB_IP, IB_PORT, clientId=98)
        portfolio_items = ib_dash.portfolio(account=TARGET_ACCOUNT)
        ib_dash.disconnect()
    except Exception:
        return  # Silent fail if IBKR is closed; just proceed to menu

    stock_items = [p for p in portfolio_items if p.contract.secType == 'STK']
    if not stock_items:
        return  # Flat portfolio, skip dashboard

    print_header("REAL-TIME PORTFOLIO SNAPSHOT (IBKR MTM)")

    print(
        f"{'TICKER':<8} {'SHARES':>8} {'COST_BASIS':>12} {'CURR_PRICE':>12} {'VAL_NOW':>12} {'P&L $':>12} {'P&L %':>10}")
    print("-" * 80)

    total_cost = 0.0
    total_val = 0.0
    total_pnl = 0.0
    winners = 0

    best_ticker, best_pct = "", -999.0
    worst_ticker, worst_pct = "", 999.0

    for p in sorted(stock_items, key=lambda x: x.contract.symbol):
        tck = p.contract.symbol
        shares = p.position
        cost = p.averageCost
        curr = p.marketPrice
        val = p.marketValue
        pnl_val = p.unrealizedPNL

        pnl_pct = (curr - cost) / cost if cost > 0 else 0.0

        total_cost += (cost * shares)
        total_val += val
        total_pnl += pnl_val

        if pnl_val > 0: winners += 1

        if pnl_pct > best_pct:
            best_pct = pnl_pct
            best_ticker = tck
        if pnl_pct < worst_pct:
            worst_pct = pnl_pct
            worst_ticker = tck

        pnl_sign_usd = "+" if pnl_val > 0 else "-"
        pnl_sign_pct = "+" if pnl_pct > 0 else "-"

        print(
            f"{tck:<8} {shares:>8.2f} ${cost:>11.2f} ${curr:>11.2f} ${val:>11.2f} {pnl_sign_usd}${abs(pnl_val):>10.2f} {pnl_sign_pct}{abs(pnl_pct) * 100:>7.2f}%")

    print("-" * 80)

    period_ret = total_pnl / total_cost if total_cost > 0 else 0.0
    win_rate = winners / len(stock_items) if stock_items else 0.0

    # Retrieve Dates for Benchmarking & Annualization
    inc_date, reb_date = None, None
    if os.path.exists(TRADE_LOG_FILE):
        try:
            df_log = pd.read_excel(TRADE_LOG_FILE)
            if not df_log.empty:
                df_log['Entry_Date'] = pd.to_datetime(df_log['Entry_Date']).dt.tz_localize(None)
                inc_date = df_log['Entry_Date'].min()
                last_reb_id = df_log['Rebalance_ID'].max()
                reb_date = df_log[df_log['Rebalance_ID'] == last_reb_id]['Entry_Date'].min()
        except:
            pass

    now = pd.Timestamp.now()

    days_active = 1
    if reb_date:
        days_active = (now - reb_date).days
        if days_active < 1: days_active = 1

    ann_ret = (1 + period_ret) ** (365 / days_active) - 1

    sign_p = "+" if total_pnl > 0 else ""
    sign_pr = "+" if period_ret > 0 else ""
    sign_ar = "+" if ann_ret > 0 else ""

    print(f"📈 TOTAL EQUITY:      ${total_val:,.2f}")
    print(f"📉 PREVIOUS EQUITY:   ${total_cost:,.2f}")
    print(f"💰 PERIOD P&L:        {sign_p}${total_pnl:,.2f}")
    print(f"🚀 PERIOD RETURN:     {sign_pr}{period_ret * 100:.2f}%  (Ann: {sign_ar}{ann_ret * 100:.2f}%)")
    print(f"📊 PORTFOLIO WIN %:   {win_rate * 100:.1f}% ({winners}/{len(stock_items)})")
    print("-" * 80)
    print(f"🏆 BEST PERFORMER:    {best_ticker} ({'+' if best_pct > 0 else ''}{best_pct * 100:.2f}%)")
    print(f"💩 WORST PERFORMER:   {worst_ticker} ({'+' if worst_pct > 0 else ''}{worst_pct * 100:.2f}%)")
    print("=" * 80)

    # --- S&P 500 QUICK BENCHMARK ---
    if inc_date and reb_date:
        print("\n" + "=" * 120)
        print(f"   Period:        {reb_date.strftime('%Y-%m-%d %H:%M')} -> {now.strftime('%Y-%m-%d %H:%M')}")
        print(f"   Total Period:  {inc_date.strftime('%Y-%m-%d %H:%M')} -> {now.strftime('%Y-%m-%d %H:%M')}")
        print("-" * 120)

        try:
            st_fetch = min(inc_date, reb_date) - pd.Timedelta(days=7)
            spy = yf.download(BENCHMARK_TICKER, start=st_fetch.strftime('%Y-%m-%d'), progress=False)
            if not spy.empty:
                if isinstance(spy.columns, pd.MultiIndex):
                    spy_c = spy['Close'][BENCHMARK_TICKER]
                else:
                    spy_c = spy['Close']
                spy_c.index = spy_c.index.tz_localize(None)

                # Nearest valid close proxy
                inc_px = spy_c[spy_c.index <= inc_date].iloc[-1] if not spy_c[spy_c.index <= inc_date].empty else \
                spy_c.iloc[0]
                reb_px = spy_c[spy_c.index <= reb_date].iloc[-1] if not spy_c[spy_c.index <= reb_date].empty else \
                spy_c.iloc[0]
                curr_px = spy_c.iloc[-1]

                inc_ret = (curr_px / inc_px) - 1
                reb_ret = (curr_px / reb_px) - 1

                print(" 📅 TOTAL PERIOD (Since Inception)")
                print(f"    S&P 500 Index Return:                    {'+' if inc_ret > 0 else ''}{inc_ret * 100:.2f}%")
                print("-" * 120)
                print(" ⏱️  CURRENT PERIOD (Since Last Rebalance)")
                print(f"    S&P 500 Index Return:                    {'+' if reb_ret > 0 else ''}{reb_ret * 100:.2f}%")
                print("=" * 120)
        except Exception:
            print(" ⚠️ Benchmark data currently unavailable.")


# ==============================================================================
# PASS 1: SIGNAL GENERATION & SCRAPING
# ==============================================================================

def fetch_live_equity_preflight():
    log("Attempting to auto-fetch live account balance from IBKR (Requesting server snapshot)...")
    ib_pre = IB()
    try:
        ib_pre.connect(IB_IP, IB_PORT, clientId=99)
        vals = ib_pre.accountSummary(TARGET_ACCOUNT)

        nlv = None
        acct_currency = "USD"
        for v in vals:
            if v.tag == 'NetLiquidation':
                nlv = float(v.value)
                acct_currency = v.currency
                break

        ib_pre.disconnect()

        # --- FX CONVERSION LOGIC ---
        if nlv is not None and acct_currency and acct_currency != 'USD':
            log(f"Account base currency detected as {acct_currency}. Fetching live FX rate for USD conversion...")
            try:
                fx_rate = yf.Ticker(f"{acct_currency}USD=X").fast_info['last_price']
                converted_nlv = nlv * fx_rate
                log(f"FX Rate ({acct_currency}/USD): {fx_rate:.4f} -> Adjusted USD Equity: ${converted_nlv:,.2f}")
                return converted_nlv
            except Exception as fx_e:
                log(f"WARNING: FX conversion failed ({fx_e}). Using raw NLV.")

        return nlv
    except Exception as e:
        log(f"Auto-fetch failed. Error: {e}")
        return None


def get_tickers(universe_choice):
    tickers = []
    if TEST_MODE:
        return ["NVDA", "AAPL", "MSFT", "AMZN", "GOOGL", "META", "TSLA", "JPM", "V", "UNH",
                "PG", "HD", "MA", "LLY", "CVX", "MRK", "ABBV", "PEP", "KO", "BAC"]

    try:
        if universe_choice == "S&P 500" or universe_choice is None:
            url = 'https://en.wikipedia.org/wiki/List_of_S%26P_500_companies'
            response = session.get(url)
            response.raise_for_status()
            table = pd.read_html(io.StringIO(response.text))
            tickers = [t.replace('.', '-') for t in table[0]['Symbol'].tolist()]

        elif universe_choice == "STOXX 600":
            tickers = ["MC.PA", "OR.PA", "ASML.AS", "SAP.DE", "SIE.DE", "TTE.PA"]

        elif universe_choice == "Russell 3000":
            urls = ['https://en.wikipedia.org/wiki/List_of_S%26P_500_companies',
                    'https://en.wikipedia.org/wiki/List_of_S%26P_400_companies',
                    'https://en.wikipedia.org/wiki/List_of_S%26P_600_companies']
            temp_tickers = set()
            for u in urls:
                try:
                    resp = session.get(u)
                    tbl = pd.read_html(io.StringIO(resp.text))
                    df_idx = tbl[0]
                    col = 'Symbol' if 'Symbol' in df_idx.columns else 'Ticker symbol'
                    if col in df_idx.columns: temp_tickers.update(df_idx[col].tolist())
                except:
                    pass
            tickers = [t.replace('.', '-') for t in list(temp_tickers)]

    except Exception as e:
        log(f"Error fetching tickers: {e}")
        return []
    return tickers


def get_financial_data(ticker):
    try:
        stock = yf.Ticker(ticker)
        info = stock.info
        try:
            price = stock.fast_info['last_price']
        except:
            price = info.get('currentPrice', info.get('regularMarketPrice', 0))

        if not price or price <= 0: return None

        eps_fwd = None
        try:
            est = stock.earnings_estimate
            if est is not None and '+1y' in est.index: eps_fwd = est.loc['+1y', 'avg']
        except:
            pass

        if not eps_fwd: eps_fwd = info.get('forwardEps')
        if not eps_fwd: return None

        dates = stock.earnings_dates
        eps_ttm_street, eps_prior_street, credibility_score, has_history = 0.0, 0.0, 999.0, False

        if dates is not None and not dates.empty:
            try:
                dates.index = dates.index.tz_localize(None)
                now = pd.Timestamp.now()
                past_earnings = dates[dates.index < now].copy()
                actual_col = next((c for c in dates.columns if 'Actual' in str(c) or 'Reported' in str(c)), None)
                est_col = next((c for c in dates.columns if 'Estimate' in str(c) or 'Consensus' in str(c)), None)

                if actual_col and not past_earnings.empty:
                    past_earnings = past_earnings.sort_index(ascending=False).dropna(subset=[actual_col])
                    if len(past_earnings) >= 4:
                        last_4 = past_earnings.head(4)
                        eps_ttm_street = last_4[actual_col].sum()
                        has_history = True
                        if est_col:
                            surprises = [abs((row[actual_col] - row[est_col]) / row[est_col])
                                         for _, row in last_4.iterrows() if row[est_col] and row[est_col] != 0]
                            if surprises: credibility_score = sum(surprises) / len(surprises)
                    if len(past_earnings) >= 8: eps_prior_street = past_earnings.iloc[4:8][actual_col].sum()
            except:
                pass

        if not has_history:
            eps_ttm_street = info.get('trailingEps', 0.01)
            eps_prior_street = eps_ttm_street * 0.9

        sector = info.get('sector', 'Unknown')
        diagnosis = "⚠️ Rebound" if eps_ttm_street < eps_prior_street else "✔ Organic"
        implied_growth = (eps_fwd - eps_ttm_street) / eps_ttm_street if eps_ttm_street != 0 else 0
        pe_curr = price / eps_ttm_street if eps_ttm_street > 0 else 0
        pe_fwd = price / eps_fwd if eps_fwd > 0 else 0
        peg_fwd = pe_curr / (implied_growth * 100) if (pe_curr > 0 and implied_growth > 0.001) else 999.0

        return {'Ticker': ticker, 'Sector': sector, 'Price': price, 'PE_TTM': pe_curr,
                'EPS_Fwd': eps_fwd, 'EPS_TTM': eps_ttm_street,
                'EPS_Prior': eps_prior_street, 'Implied_Growth': implied_growth,
                'Diagnosis': diagnosis, 'PE_Forward': pe_fwd, 'PEG_Forward': peg_fwd,
                'Credibility_Pct': credibility_score}
    except:
        return None


def build_target_portfolio():
    print_header("1. UNIVERSE & STRATEGY SIGNAL GENERATION")

    print("\n1. S&P 500\n2. STOXX 600\n3. Russell 3000")
    u_in = input("Select Universe (1/2/3): ")
    universe_name = {"1": "S&P 500", "2": "STOXX 600", "3": "Russell 3000"}.get(u_in, "S&P 500")

    try:
        cred_limit = float(input("Enter Credibility Cutoff % (e.g., 20): ")) / 100.0
        peg_limit = float(input("Enter Forward PEG Cutoff (e.g., 0.6): "))
        pe_cap_limit = float(input("Enter Forward P/E Cutoff (e.g., 9): "))
        growth_limit = float(input("Enter Implied Growth Cutoff % (e.g., 25): ")) / 100.0
        only_organic = input("Only Organic Growth? (y/n): ").lower() == 'y'
    except ValueError:
        log("Invalid input. Defaulting to safe values.")
        cred_limit, peg_limit, pe_cap_limit, growth_limit, only_organic = 0.20, 0.6, 9.0, 0.25, True

    print("\n" + "-" * 80)
    print(" CAPITAL ALLOCATION (COMPOUNDING ENGINE)")
    print("-" * 80)

    live_eq = fetch_live_equity_preflight()
    if live_eq is not None:
        print(f"\n📡 SUCCESS: Detected live IBKR Net Liquidation Value: ${live_eq:,.2f}")
        eq_in = input(f"Press Enter to use ${live_eq:,.2f} for rebalancing, or type a custom amount: ")
        active_equity = float(eq_in) if eq_in.strip() else live_eq
    else:
        eq_in = input(
            f"\nEnter current Total Equity for allocation (Press Enter to use default ${BASE_EQUITY_FALLBACK:,.2f}): ")
        active_equity = float(eq_in) if eq_in.strip() else BASE_EQUITY_FALLBACK

    log(f"Target Equity set to: ${active_equity:,.2f} (Weights will be calculated based on this amount)")

    # --- INTELLIGENT LOCAL CACHING LOGIC ---
    today_str = datetime.datetime.now().strftime("%Y%m%d")
    safe_univ = universe_name.replace(" ", "").replace("&", "n")
    local_data_file = f"data_{safe_univ}_{today_str}.xlsx"

    if os.path.exists(local_data_file):
        log(f"⚡ Local cache found for today ({local_data_file}). Bypassing YFinance scrape...")
        df = pd.read_excel(local_data_file)
    else:
        tickers = get_tickers(universe_name)
        if not tickers: return

        log(f"Initiating scrape for {len(tickers)} tickers in {universe_name}...")
        results = []
        start_time = time.time()

        for i, t in enumerate(tickers, 1):
            time.sleep(random.uniform(0.1, 0.2))
            d = get_financial_data(t)
            if d: results.append(d)
            if i % 25 == 0 or i == len(tickers):
                rate = i / (time.time() - start_time) if (time.time() - start_time) > 0 else 0
                print(f"   >> Processed {i}/{len(tickers)} | Rate: {rate:.1f} tkrs/sec | Current: {t}")

        df = pd.DataFrame(results)
        if df.empty:
            log("No valid data retrieved. Yahoo Finance may be blocking your IP.")
            return

        # Export full raw universe data
        try:
            df.to_excel(local_data_file, index=False)
            log(f"✅ Full raw universe data (all variables) exported to {local_data_file} for posterity.")
        except Exception as e:
            log(f"Warning: Could not save {local_data_file}. Error: {e}")

    # --- APPLY FILTERING LOGIC ---
    mask = (df['PE_Forward'] > 0) & (df['PEG_Forward'] < peg_limit) & \
           (df['PE_Forward'] < pe_cap_limit) & (df['Implied_Growth'] > growth_limit) & \
           (df['Credibility_Pct'] < cred_limit)
    if only_organic:
        mask &= (df['Diagnosis'] == "✔ Organic")
    elif EXCLUDE_REBOUNDS:
        mask &= (df['Diagnosis'] != "⚠️ Rebound")

    valid_universe = df[mask].copy().sort_values(by='PE_Forward', ascending=True)
    qualifying_count = len(valid_universe)
    log(f"Filtering complete. {qualifying_count} companies meet strict criteria.")

    final_selection = pd.DataFrame()
    if qualifying_count > MAX_HOLDINGS:
        final_selection = valid_universe.head(MAX_HOLDINGS)
    elif qualifying_count < MIN_HOLDINGS:
        print(f"\nWARNING: Only {qualifying_count} companies meet the criteria (Minimum is {MIN_HOLDINGS}).")
        next_best_mask = (df['PE_Forward'] >= pe_cap_limit) & (df['PEG_Forward'] < peg_limit) & \
                         (df['Implied_Growth'] > growth_limit) & (df['Credibility_Pct'] < cred_limit)
        if only_organic: next_best_mask &= (df['Diagnosis'] == "✔ Organic")
        candidates = df[next_best_mask].sort_values(by='PE_Forward', ascending=True).head(
            MIN_HOLDINGS - qualifying_count)

        if not candidates.empty:
            ans = input(f"\nFill portfolio to {MIN_HOLDINGS} with next best candidates? (y/n): ")
            if ans.lower() == 'y':
                final_selection = pd.concat([valid_universe, candidates])
            else:
                final_selection = valid_universe
        else:
            final_selection = valid_universe
    else:
        final_selection = valid_universe

    # --- SCREENER EXPORT WITH SELECTION HIGHLIGHTING ---
    try:
        raw_cols = ['Ticker', 'Sector', 'Price', 'PE_TTM', 'Implied_Growth', 'Diagnosis', 'PE_Forward', 'PEG_Forward',
                    'Credibility_Pct']
        df_export = df[raw_cols].copy()

        with pd.ExcelWriter(SCREENER_OUTPUT_FILE, engine='openpyxl') as writer:
            df_export.to_excel(writer, index=False, sheet_name='All_Scraped_Data')

            workbook = writer.book
            worksheet = writer.sheets['All_Scraped_Data']

            chosen_tickers = set(final_selection['Ticker'].tolist())
            deep_green_font = Font(color="006400", bold=True)

            for row_idx in range(2, len(df_export) + 2):
                ticker_val = worksheet.cell(row=row_idx, column=1).value
                if ticker_val in chosen_tickers:
                    for col_idx in range(1, len(raw_cols) + 1):
                        worksheet.cell(row=row_idx, column=col_idx).font = deep_green_font

        log(f"✅ Filtered screener view exported to {SCREENER_OUTPUT_FILE} with portfolio selections highlighted in green.")
    except Exception as e:
        log(f"Warning: Could not save {SCREENER_OUTPUT_FILE}. Error: {e}")

    now_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
    weight = 1.0 / len(final_selection) if not final_selection.empty else 0
    target_val_per_stock = active_equity * weight

    holdings = {}
    for _, row in final_selection.iterrows():
        price = row['Price']
        holdings[row['Ticker']] = {
            'shares': target_val_per_stock / price if price > 0 else 0,
            'entry_price': price,
            'entry_date': now_str,
            'metrics': {
                'sector': row['Sector'],
                'pe_ttm': row['PE_TTM'],
                'pe_fwd': row['PE_Forward'],
                'growth': row['Implied_Growth'],
                'peg_fwd': row['PEG_Forward'],
                'credibility': row['Credibility_Pct']
            }
        }

    db = {"Target_Portfolio": {"equity_basis": active_equity, "equity_current": active_equity, "holdings": holdings,
                               "last_update": now_str}}
    with open(FILENAME_DB, 'w') as f:
        json.dump(db, f, indent=4)
    print(f"\n✅ Target JSON generated ({FILENAME_DB}) with {len(holdings)} holdings.")


# ==============================================================================
# PASS 2 & 3: IBKR GATEWAY EXECUTION & PERSISTENT LOGGING
# ==============================================================================

def execute_ibkr_sync():
    print_header("2 & 3. IBKR GATEWAY SYNC, EXECUTION & RECONCILIATION")

    if not os.path.exists(FILENAME_DB): return log(f"CRITICAL: {FILENAME_DB} not found.")
    with open(FILENAME_DB, 'r') as f:
        db = json.load(f)
    target_holdings = db.get("Target_Portfolio", {}).get("holdings", {})
    if not target_holdings: return log("Target portfolio is empty.")

    # Validate/Initialize df_log shape early
    cols = ['Rebalance_ID', 'Portfolio', 'Ticker', 'Sector', 'Entry_Date', 'Exit_Date', 'Shares', 'Entry_Price',
            'Exit_Price',
            'P&L', 'Return_Pct', 'PE_TTM_At_Entry', 'Forward_PE_At_Entry', 'Implied_Growth_At_Entry', 'PEG_At_Entry',
            'Credibility_At_Entry']
    df_log = pd.read_excel(TRADE_LOG_FILE) if os.path.exists(TRADE_LOG_FILE) else pd.DataFrame(columns=cols)
    for c in cols:
        if c not in df_log.columns: df_log[c] = np.nan

    ib = IB()
    try:
        ib.connect(IB_IP, IB_PORT, clientId=IB_CLIENT_ID)
    except Exception as e:
        return log(f"Failed to connect to IBKR: {e}")

    if TARGET_ACCOUNT not in ib.managedAccounts(): return log(f"Account {TARGET_ACCOUNT} not found.")

    log("Retrieving filled positions and pending orders...")
    ib_positions = ib.positions(account=TARGET_ACCOUNT)
    filled_portfolio = {pos.contract.symbol: pos.position for pos in ib_positions if pos.contract.secType == 'STK'}

    ib.reqAllOpenOrders()
    ib.sleep(1)
    pending_portfolio = {}
    for t in ib.trades():
        if t.contract.secType == 'STK' and t.orderStatus.status in ['PreSubmitted', 'Submitted', 'PendingSubmit']:
            sym, rem = t.contract.symbol, t.order.totalQuantity - t.orderStatus.filled
            pending_portfolio[sym] = pending_portfolio.get(sym, 0.0) + (rem if t.order.action == 'BUY' else -rem)

    log("Calculating trade deltas...")
    all_tickers = set(list(target_holdings.keys()) + list(filled_portfolio.keys()) + list(pending_portfolio.keys()))
    delta_rows, trades_to_execute = [], []

    for ticker in all_tickers:
        t_shares = int(round(target_holdings.get(ticker, {}).get('shares', 0.0)))
        e_shares = filled_portfolio.get(ticker, 0.0) + pending_portfolio.get(ticker, 0.0)
        delta = t_shares - e_shares
        if delta != 0: trades_to_execute.append(
            {'ticker': ticker, 'action': "BUY" if delta > 0 else "SELL", 'quantity': abs(delta)})
        delta_rows.append({'Ticker': ticker, 'Filled Shares': filled_portfolio.get(ticker, 0.0),
                           'Pending Shares': pending_portfolio.get(ticker, 0.0), 'Target': t_shares, 'Delta': delta})

    active_trade_objects = []

    # Check if a rebalance is needed. If not, self-repair empty log if necessary.
    if not trades_to_execute:
        if df_log.empty and len(filled_portfolio) > 0:
            log("Detected active portfolio with an empty trade log. Bypassing execution to generate Rebalance #1 snapshot...")
        else:
            log("No trades necessary. Portfolio synced.")
            return ib.disconnect()
    else:
        df_delta = pd.DataFrame(delta_rows).sort_values(by='Delta', ascending=False)
        print("\n--- PORTFOLIO DELTA CALCULATION ---")
        print(df_delta.to_string(index=False))

        if input(f"\nFIRE {len(trades_to_execute)} ORDERS? (y/n): ").lower() != 'y': return ib.disconnect()

        # --- DYNAMIC MARKET HOURS LOGIC ---
        now_est = pd.Timestamp.now(tz='US/Eastern')
        market_open = now_est.replace(hour=9, minute=30, second=0, microsecond=0)
        market_close = now_est.replace(hour=16, minute=0, second=0, microsecond=0)
        is_open = (now_est.weekday() < 5) and (market_open <= now_est <= market_close)

        if not is_open:
            log("US Markets are currently CLOSED. Orders will be routed as Market-On-Open (MOO) with TIF='OPG'.")

        log("Routing orders...")
        for trade in trades_to_execute:
            contract = Stock(trade['ticker'], 'SMART', 'USD')
            try:
                ib.qualifyContracts(contract)
            except Exception as e:
                log(f"Contract qualification failed for {trade['ticker']}: {e}")
                continue

            if not is_open:
                order = MarketOrder(trade['action'], trade['quantity'], tif='OPG')
            else:
                order = MarketOrder(trade['action'], trade['quantity'])

            placed_trade = ib.placeOrder(contract, order)
            active_trade_objects.append({'ticker': trade['ticker'], 'trade': placed_trade})
            ib.sleep(0.5)

        log("All orders submitted to IBKR matching engine.")
        log("Waiting up to 30 seconds for fills and commission reports...")
        timeout = time.time() + 30
        while ib.waitOnUpdate(timeout=1):
            if all(t['trade'].orderStatus.status in ['Filled', 'Cancelled'] for t in
                   active_trade_objects) or time.time() > timeout: break
        ib.sleep(2)

        # --- CONSOLE RECONCILIATION REPORT ---
        print("\n" + "-" * 80)
        print(" 🔍 RECONCILIATION REPORT")
        print("-" * 80)
        for t_data in active_trade_objects:
            tck, obj = t_data['ticker'], t_data['trade']
            status = obj.orderStatus.status
            req = obj.order.totalQuantity
            filled = obj.orderStatus.filled
            avgPx = obj.orderStatus.avgFillPrice

            # Safely extract commission
            fee = 0.0
            if obj.fills:
                for fill in obj.fills:
                    if hasattr(fill, 'commissionReport') and fill.commissionReport is not None:
                        fee += getattr(fill.commissionReport, 'commission', 0.0)

            time_str = datetime.datetime.now().strftime("%H:%M:%S")
            print(
                f"[{time_str}] Result: {obj.order.action} {tck} | Req: {req}, Filled: {filled} | AvgPx: ${avgPx:.2f} | Fee: ${fee:.2f} ({status})")
        print("-" * 80)

    log("Fetching final post-execution portfolio state...")
    ib_positions_final = ib.positions(account=TARGET_ACCOUNT)
    final_portfolio = {pos.contract.symbol: pos.position for pos in ib_positions_final if pos.contract.secType == 'STK'}
    with open(REAL_FILENAME_DB, 'w') as f:
        json.dump(final_portfolio, f, indent=4)

    # Safely extract total commission
    total_comm = 0.0
    for t_data in active_trade_objects:
        if t_data['trade'].fills:
            for fill in t_data['trade'].fills:
                if hasattr(fill, 'commissionReport') and fill.commissionReport is not None:
                    total_comm += getattr(fill.commissionReport, 'commission', 0.0)

    print(f"\n[{datetime.datetime.now().strftime('%H:%M:%S')}] Total Commission Paid: ${total_comm:.2f}")
    print(f"[{datetime.datetime.now().strftime('%H:%M:%S')}] ✅ Real IBKR Portfolio state saved to: {REAL_FILENAME_DB}")
    print("=" * 80 + "\n")

    # ==============================================================================
    # THE REBALANCE LEDGER SNAPSHOT ENGINE
    # ==============================================================================
    log("Constructing Immutable Ledger Snapshot...")
    now_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    if df_log.empty:
        curr_reb_id = 1
    else:
        curr_reb_id = df_log['Rebalance_ID'].max() + 1

    # 1. Map dynamic fill prices if available
    fill_prices = {}
    for t_data in active_trade_objects:
        if t_data['trade'].orderStatus.filled > 0:
            fill_prices[t_data['ticker']] = t_data['trade'].orderStatus.avgFillPrice

    # 2. Close out ALL open positions from previous rebalance period
    open_mask = df_log['Exit_Date'].isna()
    if not df_log.empty and open_mask.any():
        for idx, row in df_log[open_mask].iterrows():
            tck = row['Ticker']
            exit_px = fill_prices.get(tck, target_holdings.get(tck, {}).get('entry_price', np.nan))

            if pd.isna(exit_px) or exit_px == 0:
                try:
                    exit_px = yf.Ticker(tck).fast_info['last_price']
                except:
                    pass

            df_log.at[idx, 'Exit_Date'] = now_str
            df_log.at[idx, 'Exit_Price'] = exit_px
            e_px = df_log.at[idx, 'Entry_Price']
            shares = df_log.at[idx, 'Shares']
            if pd.notna(e_px) and e_px > 0 and pd.notna(exit_px):
                df_log.at[idx, 'P&L'] = (exit_px - e_px) * shares
                df_log.at[idx, 'Return_Pct'] = (exit_px / e_px) - 1

    # 3. Create brand new snapshot for the incoming portfolio period
    new_rows = []
    for tck, t_data in target_holdings.items():
        ent_px = fill_prices.get(tck, t_data.get('entry_price', 0))
        metrics = t_data.get('metrics', {})

        # Ensure we log target shares even if MOO order hasn't filled yet
        actual_shares = final_portfolio.get(tck, t_data.get('shares', 0))
        if actual_shares == 0: actual_shares = t_data.get('shares', 0)

        new_rows.append({
            'Rebalance_ID': curr_reb_id,
            'Portfolio': 'Forward_PE_Base',
            'Ticker': tck,
            'Sector': metrics.get('sector', 'Unknown'),
            'Entry_Date': now_str,
            'Shares': actual_shares,
            'Entry_Price': ent_px,
            'PE_TTM_At_Entry': metrics.get('pe_ttm', np.nan),
            'Forward_PE_At_Entry': metrics.get('pe_fwd', np.nan),
            'Implied_Growth_At_Entry': metrics.get('growth', np.nan),
            'PEG_At_Entry': metrics.get('peg_fwd', np.nan),
            'Credibility_At_Entry': metrics.get('credibility', np.nan)
        })

    df_log = pd.concat([df_log, pd.DataFrame(new_rows)], ignore_index=True)
    try:
        df_log.to_excel(TRADE_LOG_FILE, index=False)
        log(f"✅ Portfolio Snapshot (Rebalance #{curr_reb_id}) permanently carved into {TRADE_LOG_FILE}.")
    except Exception as e:
        log(f"CRITICAL: Failed to write {TRADE_LOG_FILE}. Ensure it's closed.")

    ib.disconnect()


# ==============================================================================
# HELPER: STRATEGY EVALUATION METRICS
# ==============================================================================
def calc_advanced_metrics(daily_returns, rf_rate=RISK_FREE_RATE):
    if len(daily_returns) < 2: return {}

    compounded = (1 + daily_returns).cumprod()
    cagr = (compounded.iloc[-1]) ** (252 / len(daily_returns)) - 1 if len(daily_returns) > 0 else 0

    roll_max = compounded.cummax()
    drawdown = (compounded - roll_max) / roll_max
    max_dd = drawdown.min()

    vol = daily_returns.std() * np.sqrt(252)

    excess_ret = daily_returns.mean() * 252 - rf_rate
    sharpe = excess_ret / vol if vol > 0 else 0

    downside_returns = daily_returns[daily_returns < 0]
    downside_vol = downside_returns.std() * np.sqrt(252)
    sortino = excess_ret / downside_vol if downside_vol > 0 else 0

    calmar = cagr / abs(max_dd) if max_dd < 0 else 0

    return {
        'CAGR': f"{cagr:.2%}",
        'Max Drawdown': f"{max_dd:.2%}",
        'Annual Volatility': f"{vol:.2%}",
        'Sharpe Ratio': round(sharpe, 2),
        'Sortino Ratio': round(sortino, 2),
        'Calmar Ratio': round(calmar, 2)
    }


# ==============================================================================
# PASS 4: COMPLETE AUDIT REWRITE & EXCEL REPORTING
# ==============================================================================

def run_real_portfolio_audit():
    print_header("4. STATISTICAL AUDIT & MONTE CARLO REPORTING")

    if not os.path.exists(TRADE_LOG_FILE): return log(
        f"CRITICAL: {TRADE_LOG_FILE} not found. Complete an execution first.")

    df_log = pd.read_excel(TRADE_LOG_FILE)
    if df_log.empty: return log("Trade log is empty. No metrics to run.")

    # Standardize Dates securely - Normalize strips the time to 00:00:00 for strict daily alignment
    df_log['Entry_Date'] = pd.to_datetime(df_log['Entry_Date']).dt.tz_localize(None).dt.normalize()
    df_log['Exit_Date'] = pd.to_datetime(df_log['Exit_Date']).dt.tz_localize(None).dt.normalize()

    inception_date = df_log['Entry_Date'].min()
    last_reb_id = df_log['Rebalance_ID'].max()
    last_reb_date = df_log[df_log['Rebalance_ID'] == last_reb_id]['Entry_Date'].min()
    today = pd.Timestamp.now().replace(tzinfo=None)

    log("Downloading SP500 constituents and benchmark for equal-weight construction...")
    sp500_tickers = get_tickers("S&P 500")
    all_tickers = list(set(df_log['Ticker'].dropna().tolist() + sp500_tickers + [BENCHMARK_TICKER]))

    start_fetch = min(inception_date, last_reb_date) - pd.Timedelta(days=7)  # Buffer for weekend starts
    data = yf.download(all_tickers, start=start_fetch.strftime('%Y-%m-%d'), progress=False, auto_adjust=True)

    if isinstance(data.columns, pd.MultiIndex):
        prices = data['Close'] if 'Close' in data.columns.levels[0] else data.xs(data.columns.levels[0][0], level=0,
                                                                                 axis=1)
    else:
        prices = data

    prices = prices.ffill()
    prices.index = prices.index.tz_localize(None)
    pct_returns = prices.pct_change().fillna(0)

    bench_cap = pct_returns[BENCHMARK_TICKER] if BENCHMARK_TICKER in pct_returns.columns else pd.Series(0,
                                                                                                        index=pct_returns.index)
    valid_sp500 = [t for t in sp500_tickers if t in pct_returns.columns]
    bench_ew = pct_returns[valid_sp500].mean(axis=1)

    today_str = today.strftime("%Y%m%d")
    report_file = f"audit_report_{today_str}.xlsx"

    try:
        writer = pd.ExcelWriter(report_file, engine='openpyxl')
    except Exception as e:
        return log(f"CRITICAL: Cannot open Excel Writer. Is {report_file} currently open? Error: {e}")

    periods = {
        "Inception": {"start": inception_date, "log": df_log},
        "LastRebalance": {"start": last_reb_date, "log": df_log[df_log['Rebalance_ID'] == last_reb_id]}
    }

    log("Running Statistical Engines & Generating Visualizations...")

    for p_name, p_data in periods.items():
        st_date = p_data["start"]
        sub_log = p_data["log"]

        mask = (prices.index >= st_date)
        p_prices, p_pct = prices.loc[mask], pct_returns.loc[mask]
        p_b_cap, p_b_ew = bench_cap.loc[mask], bench_ew.loc[mask]

        if p_prices.empty: continue

        # --- BLOCK 1: STRATEGY DAILY RETURNS & VISUALIZATIONS ---
        log(f"[{p_name}] Calculating timeseries and generating charts...")
        strat_daily = pd.Series(0.0, index=p_prices.index)
        strat_breadth = pd.Series(0.0, index=p_prices.index)

        for date in p_prices.index:
            active = sub_log[(sub_log['Entry_Date'] <= date) &
                             (pd.isna(sub_log['Exit_Date']) | (sub_log['Exit_Date'] > date))]
            active_tkrs = [t for t in active['Ticker'].tolist() if t in p_pct.columns]

            if active_tkrs:
                strat_daily.loc[date] = p_pct.loc[date, active_tkrs].mean()

            if not active.empty:
                gains = 0
                for _, row in active.iterrows():
                    tkr = row['Ticker']
                    if tkr in p_prices.columns:
                        if p_prices.loc[date, tkr] > row['Entry_Price']: gains += 1
                strat_breadth.loc[date] = (gains / len(active)) * 100

        # Cumulative Calculations
        strat_cum = (1 + strat_daily).cumprod()
        cap_cum = (1 + p_b_cap).cumprod()
        ew_cum = (1 + p_b_ew).cumprod()

        strat_tot = strat_cum.iloc[-1] - 1
        cap_tot = cap_cum.iloc[-1] - 1
        ew_tot = ew_cum.iloc[-1] - 1

        sp500_univ_rets = (p_prices[valid_sp500].iloc[-1] / p_prices[valid_sp500].iloc[0]) - 1
        breadth_pct = (sp500_univ_rets > 0).sum() / len(valid_sp500)

        sp500_cum_daily = (1 + p_pct[valid_sp500]).cumprod() - 1
        sp500_breadth = (sp500_cum_daily > 0).mean(axis=1) * 100

        # Win rate (Closed + MTM)
        wins, total_trades = 0, len(sub_log)
        for _, row in sub_log.iterrows():
            entry_px = row['Entry_Price']
            if pd.isna(row['Exit_Price']) and row['Ticker'] in p_prices.columns:
                exit_px = p_prices[row['Ticker']].iloc[-1]
            else:
                exit_px = row['Exit_Price']

            if pd.notna(entry_px) and pd.notna(exit_px) and entry_px > 0:
                if (exit_px / entry_px) - 1 > 0: wins += 1

        win_rate = wins / total_trades if total_trades > 0 else 0

        # Generate Chart 1: Equity Curve
        plt.figure(figsize=(10, 5))
        plt.plot(strat_cum * 100 - 100, label='Strategy', color='blue', linewidth=2)
        plt.plot(cap_cum * 100 - 100, label='S&P 500 (Cap)', color='red', alpha=0.7)
        plt.plot(ew_cum * 100 - 100, label='S&P 500 (EW)', color='green', alpha=0.7)
        plt.title(f"Equity Curve Analysis - {p_name}")
        plt.ylabel("Cumulative Return (%)")
        plt.legend()
        plt.grid(True, alpha=0.3)
        plt.tight_layout()
        eq_img_path = f"equity_{p_name}.png"
        plt.savefig(eq_img_path)
        plt.close()

        # Generate Chart 2: Market Breadth
        plt.figure(figsize=(10, 5))
        plt.plot(strat_breadth, label='Strategy (% Positive from Entry)', color='blue', linewidth=2)
        plt.plot(sp500_breadth, label='S&P 500 (% Positive from Start)', color='red', alpha=0.7)
        plt.title(f"Market Breadth Comparison - {p_name}")
        plt.ylabel("% of Stocks in Profit")
        plt.legend()
        plt.grid(True, alpha=0.3)
        plt.tight_layout()
        br_img_path = f"breadth_{p_name}.png"
        plt.savefig(br_img_path)
        plt.close()

        # Advanced Performance Metrics Table
        adv_metrics = calc_advanced_metrics(strat_daily)
        df_perf = pd.DataFrame(list(adv_metrics.items()), columns=['Metric', 'Value'])

        df_breadth = pd.DataFrame([{
            'Metric': 'Strategy Total Return', 'Value': f"{strat_tot:.2%}"},
            {'Metric': 'S&P 500 Cap-Weighted Return', 'Value': f"{cap_tot:.2%}"},
            {'Metric': 'S&P 500 Equal-Weighted Return', 'Value': f"{ew_tot:.2%}"},
            {'Metric': 'Market Breadth (SP500 > 0%)', 'Value': f"{breadth_pct:.2%}"},
            {'Metric': 'Strategy Win Rate (incl. MTM)', 'Value': f"{win_rate:.2%}"}
        ])
        df_summary = pd.concat([df_breadth, df_perf], ignore_index=True)
        df_summary.to_excel(writer, sheet_name=f'Summary_{p_name}', index=False)

        # Year-by-Year Table Breakdown
        yby = strat_daily.groupby(strat_daily.index.year).apply(lambda x: (1 + x).cumprod().iloc[-1] - 1)
        yby_b = p_b_cap.groupby(p_b_cap.index.year).apply(lambda x: (1 + x).cumprod().iloc[-1] - 1)
        df_yby = pd.DataFrame({'Strategy': yby, 'Benchmark': yby_b})
        df_yby['Diff'] = df_yby['Strategy'] - df_yby['Benchmark']
        df_yby = df_yby.applymap(lambda x: f"{x:.2%}")
        df_yby.to_excel(writer, sheet_name=f'YearByYear_{p_name}')

        # --- BLOCK 2: ALPHA STATS ---
        log(f"[{p_name}] Running Regressions and T-Tests...")
        stats_rows = []
        for b_name, b_series in [('Cap-Weighted', p_b_cap), ('Equal-Weighted', p_b_ew)]:
            excess = strat_daily - b_series
            t_stat, p_val = stats.ttest_1samp(excess, 0, nan_policy='omit')
            ann_alpha = excess.mean() * 252

            ols = stats.linregress(b_series, strat_daily)
            warning = "⚠️" if pd.notna(p_val) and p_val > 0.05 else "✅"

            stats_rows.append({
                'Benchmark': b_name,
                'Daily Mean Excess': f"{excess.mean():.4%}",
                'Annualized Alpha': f"{ann_alpha:.2%}",
                'T-Stat': round(t_stat, 2) if pd.notna(t_stat) else np.nan,
                'P-Value': round(p_val, 4) if pd.notna(p_val) else np.nan,
                'Significance': warning,
                'OLS Jensen Alpha (Daily)': f"{ols.intercept:.4%}",
                'OLS Beta': round(ols.slope, 2),
                'OLS P-Value': round(ols.pvalue, 4),
                'Obs (Days)': len(excess)
            })
        pd.DataFrame(stats_rows).to_excel(writer, sheet_name=f'Stats_{p_name}', index=False)

        # --- BLOCK 3: MONTE CARLO ---
        log(f"[{p_name}] Executing 10,000 Monte Carlo simulations...")
        clean_u_rets = sp500_univ_rets.dropna().tolist()

        # Determine holdings based on timeframe constraints
        if p_name == "Inception":
            mc_holdings = 25
        else:
            mc_holdings = len(sub_log['Ticker'].unique())

        if mc_holdings > 0 and len(clean_u_rets) >= mc_holdings:
            sim_results = np.array(
                [np.random.choice(clean_u_rets, mc_holdings, replace=False).mean() for _ in range(MONTE_CARLO_SIMS)])
            n_beaten = np.sum(strat_tot > sim_results)
            mc_pval = 1.0 - (n_beaten / MONTE_CARLO_SIMS)
            mc_verdict = "PASS ✅" if mc_pval < 0.05 else "FAIL ❌"

            plt.figure(figsize=(10, 5))
            plt.hist(sim_results * 100, bins=50, color='gray', alpha=0.6, label='Random Portfolios')
            plt.axvline(strat_tot * 100, color='red', linewidth=3, label=f'Strategy ({strat_tot:.2%})')
            plt.axvline(np.mean(sim_results) * 100, color='blue', linestyle='--',
                        label=f'Random Mean ({np.mean(sim_results):.2%})')
            plt.title(
                f"Monte Carlo Distribution - {p_name}\nHoldings Assessed={mc_holdings}, Sims={MONTE_CARLO_SIMS}\nP-Val: {mc_pval:.4f} | {mc_verdict}")
            plt.xlabel("Total Return (%)")
            plt.legend()
            plt.tight_layout()

            mc_img_path = f"mc_{p_name}.png"
            plt.savefig(mc_img_path)
            plt.close()

            df_mc = pd.DataFrame([{
                'Simulations': MONTE_CARLO_SIMS,
                'Holdings Assessed': mc_holdings,
                'Target Return': f"{strat_tot:.2%}",
                'Random Avg': f"{np.mean(sim_results):.2%}",
                'P-Value': mc_pval,
                'Verdict': mc_verdict
            }])
            df_mc.to_excel(writer, sheet_name=f'MonteCarlo_{p_name}', index=False)

            # Package all visuals securely into one clean sheet
            ws_visuals = writer.book.create_sheet(f'Visuals_{p_name}')
            img1 = OpenpyxlImage(eq_img_path)
            img2 = OpenpyxlImage(br_img_path)
            img3 = OpenpyxlImage(mc_img_path)

            ws_visuals.add_image(img1, 'A1')
            ws_visuals.add_image(img2, 'A28')
            ws_visuals.add_image(img3, 'A55')

    # Output State Transition log
    df_log.to_excel(writer, sheet_name='Trade_Log_Raw', index=False)
    writer.close()

    # Cleanup pngs
    for p in ["Inception", "LastRebalance"]:
        for prefix in ["mc", "equity", "breadth"]:
            f = f"{prefix}_{p}.png"
            if os.path.exists(f): os.remove(f)

    log(f"✅ Full Statistical Audit Complete. Report saved to: {report_file}")


# ==============================================================================
# MAIN CONTROL FLOW
# ==============================================================================
if __name__ == "__main__":
    print_header("FORWARD P/E LIVE TRADING MONOLITH")
    display_startup_dashboard()

    while True:
        print("\n1. Run Screener & Generate Target JSON")
        print("2. Connect to IBKR, Execute & Reconcile")
        print("3. Audit Real Executed Portfolio")
        print("4. Run Full End-to-End Sequence")
        print("5. Exit")

        choice = input("\nSelect an option (1-5): ")

        if choice == '1':
            build_target_portfolio()
        elif choice == '2':
            execute_ibkr_sync()
        elif choice == '3':
            run_real_portfolio_audit()
        elif choice == '4':
            build_target_portfolio()
            execute_ibkr_sync()
            run_real_portfolio_audit()
        elif choice == '5':
            break
        else:
            print("Invalid choice.")